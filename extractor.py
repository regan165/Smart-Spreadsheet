import pandas as pd
from openpyxl import load_workbook

class Table:
    def __init__(self):
        self.top = 0
        self.left = 0
        self.right = 0
        self.bottom = 0
        self.name = ''

    def to_text(self, sheet):
        lines = []
        for i, row in enumerate(self.data):
            lines.append('|'.join(str(e) for e in self.data[i]))
        return "\r\n".join(lines)

class ExcelTableExtractor:
    def __init__(self):
        self.tables = []

    def open_file(self, path):
        workbook = load_workbook(filename=path, data_only=True)
        return workbook

    def has_border(self, cell):
        return cell.border.top.style or cell.border.bottom.style or cell.border.left.style or cell.border.right.style

    def is_table_header(self, cell):
        return cell.font.bold or self.has_border(cell)

    def extract_tables(self, sheet):
        tables = []
        current_table = []
        headers_detected = False

        for row in sheet.iter_rows():
            if any(self.has_border(cell) for cell in row):
                if not headers_detected:
                    headers_detected = True
                    current_table = [[cell.value for cell in row]]
                else:
                    current_table.append([cell.value for cell in row])
            else:
                if headers_detected:
                    tables.append(current_table)
                    headers_detected = False
                    current_table = []

        if headers_detected:
            tables.append(current_table)

        return tables

    def to_dataframes(self, tables):
        dataframes = []
        for table in tables:
            df = pd.DataFrame(table[1:], columns=table[0])
            dataframes.append(df)
        return dataframes

    def clean_csv_content(self, csv_content):
        lines = csv_content.split('\n')
        cleaned_lines = []
        for line in lines:
            cleaned_line = line.strip(',')
            if cleaned_line:
                cleaned_lines.append(cleaned_line)
        return '\n'.join(cleaned_lines)
